import openpyxl
import random
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, CallbackQueryHandler, ContextTypes
from config import TOKEN

EXCEL_FILE = "logs_structured.xlsx"
pending_payments = {}  # chat_id: row

def get_available_brands():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    brands = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row[3:]):
            brands.add(row[0])
    wb.close()
    return sorted(list(brands))

def get_offers_by_brand(brand):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    offers = []
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value == brand and any(cell.value for cell in row[3:]):
            offers.append({
                "row": i,
                "type_commande": row[1].value,
                "prix": row[2].value
            })
    wb.close()
    return offers

def get_random_log_and_delete(row_number):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    row = sheet[row_number]
    log_cells = row[3:503]
    available_logs = [cell for cell in log_cells if cell.value]
    if not available_logs:
        wb.close()
        return None
    chosen_cell = random.choice(available_logs)
    log = chosen_cell.value
    chosen_cell.value = None
    wb.save(EXCEL_FILE)
    wb.close()
    return log

def generate_paypal_link(chat_id, row, brand, type_commande, price):
    item_name = f"{brand} {type_commande}".replace(" ", "+")
    return (
        f"https://www.paypal.com/cgi-bin/webscr?cmd=_xclick"
        f"&business=lucasbruges4@gmail.com"
        f"&item_name={item_name}"
        f"&amount={price:.2f}"
        f"&currency_code=EUR"
        f"&notify_url=https://bot-production-608c.up.railway.app/paypal-ipn"
        f"&custom={chat_id}_{row}"
    )

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    brands = get_available_brands()
    if not brands:
        await update.message.reply_text("Aucune marque disponible.")
        return
    keyboard = [[InlineKeyboardButton(brand, callback_data=f"brand_{brand}")] for brand in brands]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("📦 Choisis une marque :", reply_markup=reply_markup)

async def handle_brand(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    brand = query.data.split("_")[1]
    offers = get_offers_by_brand(brand)
    if not offers:
        await query.edit_message_text("Aucune offre disponible pour cette marque.")
        return
    keyboard = [
        [InlineKeyboardButton(f"{offer['type_commande']} | {offer['prix']}", callback_data=f"select_{offer['row']}")]
        for offer in offers
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(f"🛍️ Offres pour {brand} :", reply_markup=reply_markup)

async def handle_select_offer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    row = int(query.data.split("_")[1])
    chat_id = query.from_user.id
    pending_payments[chat_id] = row

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    row_data = sheet[row]
    brand = row_data[0].value
    type_commande = row_data[1].value
    price = float(row_data[2].value.replace("€", "").strip()) if isinstance(row_data[2].value, str) else row_data[2].value
    wb.close()

    paypal_url = generate_paypal_link(chat_id, row, brand, type_commande, price)

    keyboard = [
        [InlineKeyboardButton("💳 Payer par PayPal", url=paypal_url)],
        [InlineKeyboardButton("🪙 Payer en Crypto (LTC)", url="https://tonsitecrypto.com/payer")],
        [InlineKeyboardButton("📱 Virement (instantané uniquement)", callback_data=f"virement_{row}")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text("💰 Choisis un moyen de paiement :", reply_markup=reply_markup)

async def handle_virement(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "📱 Pour le virement instantané, utilise ce RIB :\n"
        "`FR76 4061 8803 8800 0402 4615 147`\n\n"
        "⚠️ Contacte le SAV ici 👉 [@Razmoquette2](https://t.me/Razmoquette2) pour prouver ton paiement.\n"
        "L'envoi du log sera fait manuellement après vérification.",
        parse_mode="Markdown"
    )

app = ApplicationBuilder().token(TOKEN).build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CallbackQueryHandler(handle_brand, pattern="^brand_"))
app.add_handler(CallbackQueryHandler(handle_select_offer, pattern="^select_"))
app.add_handler(CallbackQueryHandler(handle_virement, pattern="^virement_"))

if __name__ == '__main__':
    app.run_polling()
